from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data import config
from utils.db_api.database import db


class Export:
    @staticmethod
    async def wallets() -> None:
        wallets = list(db.execute('SELECT * FROM wallets'))

        if wallets:
            spreadsheet = Workbook()
            sheet: Worksheet = spreadsheet['Sheet']
            for column, header in enumerate(['n'] + list(db.execute('SELECT * FROM wallets').keys())[1:]):
                sheet.cell(row=1, column=column + 1).value = header

            for row, wallet in enumerate(wallets):
                sheet.cell(row=row + 2, column=1).value = row + 1
                for column, value in enumerate(wallet[1:]):
                    sheet.cell(row=row + 2, column=column + 2).value = value

            spreadsheet.save(config.EXPORT_FILE)
            print(f'Done! You wallets exported to the {config.LIGHTGREEN_EX}export.xlsx{config.RESET_ALL} file.')

        else:
            print(f"{config.RED}You don't have any addresses added to the DB!{config.RESET_ALL}")
